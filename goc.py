import requests
import time
import random
import re
import threading
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

# ---------------------------------------
# Hàm lấy proxy từ API key
# ---------------------------------------
def get_proxy_from_api(api_key):
    """
    Lấy proxy từ API key thông qua API WWProxy
    Trả về chuỗi proxy nếu thành công, None nếu thất bại, hoặc dict với key "wait" nếu cần đợi
    """
    api_url = f"https://wwproxy.com/api/client/proxy/available?key={api_key}&provinceId=-1"
    
    try:
        response = requests.get(api_url)
        data = response.json()
        
        if data.get("status") == "OK" and data.get("data") and data["data"].get("proxy"):
            proxy = data["data"]["proxy"]
            print(f"Đã lấy được proxy: {proxy} từ API key: {api_key}")
            return proxy
        elif data.get("status") == "BAD_REQUEST" and "Vui lòng chờ thêm" in data.get("message", ""):
            # Trích xuất thời gian cần chờ từ thông báo lỗi
            wait_time_match = re.search(r'(\d+)s', data.get("message", ""))
            if wait_time_match:
                wait_seconds = int(wait_time_match.group(1)) + 2
                print(f"Cần đợi {wait_seconds} giây trước khi lấy proxy mới từ API key: {api_key}")
                return {"wait": wait_seconds}
        
        print(f"Lỗi khi lấy proxy từ API key {api_key}: {data}")
        return None
    except Exception as e:
        print(f"Exception khi lấy proxy từ API key {api_key}: {e}")
        return None

# ---------------------------------------
# Hàm cập nhật proxy cho profile
# ---------------------------------------
def update_proxy(profile_id, raw_proxy):
    """
    Cập nhật proxy cho profile trước khi mở.
    Gửi POST request với raw_proxy.
    Nếu trả về "Profile not found" thì log profile_id ra file profileloi.txt.
    """
    update_url = f"http://127.0.0.1:19995/api/v3/profiles/update/{profile_id}"
    headers = {"accept": "application/json", "Content-Type": "application/json"}
    data = {"raw_proxy": f"{raw_proxy}"} # Không cần thêm http:// prefix
    try:
        r = requests.post(update_url, headers=headers, json=data)
        r.raise_for_status()
        response_json = r.json()
        if response_json.get("success"):
            print(f"Proxy updated successfully for profile {profile_id}.")
            return True
        elif response_json.get("message") == "Profile not found":
            print(f"Update failed. Profile not found: {profile_id}")
            with open("profileloi.txt", "a") as f:
                f.write(str(profile_id) + "\n")
            return False
        else:
            print(f"Unexpected response when updating proxy for profile {profile_id}: {response_json}")
            return False
    except Exception as e:
        print(f"Exception updating proxy for profile {profile_id}: {e}")
        return False

# ---------------------------------------
# Đọc file proxy.txt - Đọc API keys
# ---------------------------------------
try:
    with open("proxy.txt", "r") as f:
        api_keys = [line.strip() for line in f.readlines()]
    
    if len(api_keys) < 6:
        print(f"Cảnh báo: Có ít hơn 6 API key trong file proxy.txt. Một số luồng sẽ không hoạt động.")
    
    print(f"Đã tải {len(api_keys)} API key từ file proxy.txt")

except Exception as e:
    print(f"Error reading proxy.txt: {e}")
    api_keys = []

# ---------------------------------------
# Đọc file Excel profiles.xlsx
#  - Cột A: Profile ID
#  - Cột F: Kết quả (có thể chứa "Thành Công")
# Chỉ lấy profile nếu cột F là "Thành Công"
# ---------------------------------------
workbook = load_workbook('profiles.xlsx')
worksheet = workbook.active

red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

profiles = []
for i, row in enumerate(worksheet.iter_rows(min_row=2, max_col=6, values_only=False), start=2):
    profile_id = row[0].value  # Cột A
    result_value = worksheet.cell(row=i, column=6).value  # Cột F

    # Chỉ xử lý nếu cột F là "Thành Công"
    if result_value and str(result_value).strip() == "Thành Công":
        profiles.append({
            "id": profile_id,
            "row": i
        })
    else:
        print(f"Skipping profile {profile_id} (row {i}) vì cột F không phải 'Thành Công'.")

# Biến toàn cục để xử lý đa luồng
profiles_lock = threading.Lock()
profile_index = 0  # index profile hiện tại

def process_profile(thread_id, api_key, window_pos):
    global profile_index

    while True:
        with profiles_lock:
            if profile_index >= len(profiles):
                print(f"Thread {thread_id}: No more profiles to process.")
                break
            current_profile_index = profile_index
            profile_index += 1

        profile = profiles[current_profile_index]
        profile_id = profile["id"]
        row_number = profile["row"]

        print(f"Thread {thread_id}: Processing profile {profile_id} (Row {row_number})")

        # 1) Lấy proxy từ API key và cập nhật proxy cho profile
        proxy = None
        while True:
            proxy_result = get_proxy_from_api(api_key)
            
            if isinstance(proxy_result, dict) and "wait" in proxy_result:
                # Cần đợi trước khi lấy proxy mới
                wait_seconds = proxy_result["wait"]
                print(f"Thread {thread_id}: Đợi {wait_seconds} giây trước khi lấy proxy mới.")
                time.sleep(wait_seconds)
                continue
            elif proxy_result:
                # Đã lấy được proxy
                proxy = proxy_result
                break
            else:
                # Lỗi khi lấy proxy, thử lại sau 10 giây
                print(f"Thread {thread_id}: Lỗi khi lấy proxy, thử lại sau 10 giây.")
                time.sleep(10)
                continue
                
        # Cập nhật proxy cho profile
        if not update_proxy(profile_id, proxy):
            print(f"Thread {thread_id}: Skipping profile {profile_id} due to proxy update failure.")
            continue

        # 3) Mở profile qua API
        start_url = f"http://127.0.0.1:19995/api/v3/profiles/start/{profile_id}?addination_args=--lang%3Dvi&win_pos={window_pos}&win_size=1800%2C1080&win_scale=0.35"
        print(f"Thread {thread_id}: Opening profile via URL: {start_url}")
        try:
            start_resp = requests.get(start_url)
            start_resp.raise_for_status()
        except Exception as e:
            print(f"Thread {thread_id}: Error opening profile {profile_id}: {e}")
            continue

        start_data = start_resp.json()
        if not start_data.get("success"):
            print(f"Thread {thread_id}: Failed to open profile {profile_id}: {start_data}")
            continue

        driver_path = start_data.get("data", {}).get("driver_path")
        remote_debugging_address = start_data.get("data", {}).get("remote_debugging_address")
        browser_location = start_data.get("data", {}).get("browser_location")

        if not driver_path or not remote_debugging_address:
            print(f"Thread {thread_id}: Missing driver_path or remote_debugging_address, skipping profile.")
            continue

        # 4) Khởi tạo Selenium
        options = Options()
        # options.binary_location = browser_location # nếu cần
        options.add_experimental_option("debuggerAddress", remote_debugging_address)
        service = Service(executable_path=driver_path)
        try:
            driver = webdriver.Chrome(service=service, options=options)
        except Exception as e:
            print(f"Thread {thread_id}: Error initializing webdriver for profile {profile_id}: {e}")
            continue

        # ---------------------------
        # 5) Truy cập trang tiktok
        # ---------------------------
        wait = WebDriverWait(driver, 30)
        tiktok_url = "https://www.tiktok.com"
        print(f"Thread {thread_id}: Navigating to {tiktok_url}")
        try:
            driver.get(tiktok_url)
            wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
            print(f"Thread {thread_id}: TikTok page loaded.")
        except Exception as e:
            print(f"Thread {thread_id}: Error loading TikTok page for profile {profile_id}: {e}")
            close_url = f"http://127.0.0.1:19995/api/v3/profiles/close/{profile_id}"
            try:
                requests.get(close_url)
                print(f"Thread {thread_id}: Profile {profile_id} closed successfully after error.")
            except Exception as err_close:
                print(f"Thread {thread_id}: Error closing profile {profile_id}: {err_close}")
            continue

        # (Tại đây, bạn có thể thêm code upload video nếu muốn)
        
        # 6) Do đã lọc sẵn cột F = "Thành Công", có thể không cần ghi Excel nữa

        # 7) Đóng profile qua API
        close_url = f"http://127.0.0.1:19995/api/v3/profiles/close/{profile_id}"
        print(f"Thread {thread_id}: Closing profile with URL: {close_url}")
        try:
            requests.get(close_url)
            print(f"Thread {thread_id}: Profile {profile_id} closed successfully.")
        except Exception as e:
            print(f"Thread {thread_id}: Error closing profile {profile_id}: {e}")
        time.sleep(1)

# ---------------------------------------
# Main
# ---------------------------------------
if __name__ == "__main__":
    # Xác định số lượng luồng dựa trên số API key có sẵn
    num_threads = min(6, len(api_keys))
    
    if num_threads == 0:
        print("Không có API key nào trong file proxy.txt. Không thể chạy chương trình.")
        exit(1)
    
    # Tạo và chạy các luồng dựa trên số API key có sẵn
    threads = []
    window_positions = ["0,0", "1800,0", "3600,0", "0,1080", "1800,1080", "3600,1080", "0,2160"]
    
    for i in range(num_threads):
        thread = threading.Thread(
            target=process_profile, 
            args=(i+1, api_keys[i], window_positions[i])
        )
        threads.append(thread)
        thread.start()
        
    # Chờ tất cả các luồng hoàn thành
    for thread in threads:
        thread.join()

    print("\nCompleted processing all profiles.")
